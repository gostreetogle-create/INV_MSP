"""COM bridge to a running Autodesk Inventor instance.

Single source of truth for Inventor automation, shared by:
- mcp_server.py (live tools for Claude during development)
- app.py (portable button panel for the end user)

Requires Inventor to already be running on this machine (GetActiveObject
attaches to the existing instance rather than launching a new one).
"""
import glob
import os
import re
import sys
import win32com.client

ILOGIC_ADDIN_GUID = "{3BDD8D79-2179-4B11-8A5A-257B1C0263AC}"

kPartDocumentObject = 12290
kAssemblyDocumentObject = 12291
kNewBodyOperation = 20485
kPositiveExtentDirection = 20993
kHorizontalDim = 19201
kVerticalDim = 19202

TEMPLATES_DIR = r"C:\Users\Public\Documents\Autodesk\Inventor 2027\Templates\ru-RU\Metric"
PART_TEMPLATE = TEMPLATES_DIR + r"\Standard (mm).ipt"
ASSEMBLY_TEMPLATE = TEMPLATES_DIR + r"\Standard (mm).iam"

DEFAULT_SAVE_DIR = r"C:\Users\User\Documents\Inventor\Inventor_MSP"

# When frozen by PyInstaller, __file__ resolves inside the temp extraction dir (_MEIPASS),
# which is wiped and re-extracted on every launch — a catalog stored there would silently
# discard the user's Excel edits between runs. Keep it next to the .exe instead in that case.
_APP_DIR = os.path.dirname(sys.executable) if getattr(sys, "frozen", False) else os.path.dirname(os.path.abspath(__file__))
CATALOG_PATH = os.path.join(_APP_DIR, "catalog.xlsx")

# PDM designation per bridge/NAMING.md: three 4-digit groups, e.g. 1000.0002.0004
DESIGNATION_RE = re.compile(r"^\d{4}\.\d{4}\.\d{4}$")
DESIGNATION_FILE_RE = re.compile(r"^(\d{4})\.(\d{4})\.(\d{4})$")

_app = None

# Inventor's COM API works in centimeters internally regardless of document display units.
def _mm(value):
    return value / 10.0


def connect():
    """Attach to the running Inventor instance. Raises if none is running."""
    global _app
    if _app is not None:
        try:
            _app.Caption  # noqa: B018 - probe that the COM object is still alive
            return _app
        except Exception:
            _app = None
    _app = win32com.client.GetActiveObject("Inventor.Application")
    return _app


def get_status():
    app = connect()
    has_docs = app.Documents.Count > 0
    return {
        "caption": app.Caption,
        "documents_open": app.Documents.Count,
        "active_document": app.ActiveDocument.DisplayName if has_docs else None,
        "active_document_type": app.ActiveDocument.DocumentType if has_docs else None,
    }


def list_documents():
    app = connect()
    return [
        {"name": d.DisplayName, "type": d.DocumentType, "full_path": d.FullFileName}
        for d in app.Documents
    ]


def _get_document(name=None):
    app = connect()
    if name is None:
        if app.Documents.Count == 0:
            raise RuntimeError("No document is open in Inventor")
        return app.ActiveDocument
    for d in app.Documents:
        if d.DisplayName == name or d.FullFileName == name:
            return d
    raise RuntimeError(f"No open document matches {name!r}")


def open_document(path):
    app = connect()
    return app.Documents.Open(path, True).DisplayName


def list_parameters(document_name=None):
    doc = _get_document(document_name)
    params = doc.ComponentDefinition.Parameters
    out = []
    for p in params:
        out.append({"name": p.Name, "value": p.Value, "expression": p.Expression, "units": p.Units})
    return out


def get_parameter(name, document_name=None):
    doc = _get_document(document_name)
    p = doc.ComponentDefinition.Parameters.Item(name)
    return {"name": p.Name, "value": p.Value, "expression": p.Expression, "units": p.Units}


def set_parameter(name, expression, document_name=None):
    doc = _get_document(document_name)
    p = doc.ComponentDefinition.Parameters.Item(name)
    p.Expression = expression
    doc.Update()
    return get_parameter(name, document_name)


def _ilogic_automation():
    app = connect()
    addin = app.ApplicationAddIns.ItemById(ILOGIC_ADDIN_GUID)
    addin.Activate()
    return addin.Automation


def list_ilogic_rules(document_name=None):
    doc = _get_document(document_name)
    auto = _ilogic_automation()
    rules = auto.Rules(doc)  # returns None (not an empty collection) when there are no rules yet
    return [r.Name for r in rules] if rules else []


def run_ilogic_rule(rule_name, document_name=None):
    doc = _get_document(document_name)
    auto = _ilogic_automation()
    auto.RunRule(doc, rule_name)
    return {"ran": rule_name, "document": doc.DisplayName}


def save_document(document_name=None):
    doc = _get_document(document_name)
    doc.Save()
    return {"saved": doc.DisplayName}


def export_document(target_path, document_name=None):
    """Export via Inventor's generic SaveAs — format is inferred from target_path's extension."""
    doc = _get_document(document_name)
    doc.SaveAs(target_path, True)
    return {"exported_to": target_path}


def _validate_designation(designation):
    if not DESIGNATION_RE.match(designation):
        raise ValueError(
            f"Обозначение {designation!r} не соответствует формату 0000.0000.0000 (см. NAMING.md)"
        )


def _apply_pdm_properties(doc, designation, description, material_name):
    _validate_designation(designation)
    dt = doc.PropertySets.Item("Design Tracking Properties")
    dt.Item("Part Number").Value = designation
    if description:
        dt.Item("Description").Value = description
    if material_name:
        for m in doc.Materials:
            if m.Name == material_name:
                doc.ComponentDefinition.Material = m
                break
        else:
            raise ValueError(f"Материал {material_name!r} не найден в библиотеке документа")


def _scan_designations(save_dir=None):
    """(a, b, c) int tuples for every 0000.0000.0000.(ipt|iam) file already in save_dir."""
    save_dir = save_dir or DEFAULT_SAVE_DIR
    out = []
    for path in glob.glob(os.path.join(save_dir, "*.ipt")) + glob.glob(os.path.join(save_dir, "*.iam")):
        base = os.path.splitext(os.path.basename(path))[0]
        m = DESIGNATION_FILE_RE.match(base)
        if m:
            out.append(tuple(int(x) for x in m.groups()))
    return out


def designation_exists(designation, save_dir=None):
    save_dir = save_dir or DEFAULT_SAVE_DIR
    return any(
        os.path.exists(os.path.join(save_dir, designation + ext)) for ext in (".ipt", ".iam")
    )


def _check_not_duplicate(designation, save_dir):
    if designation_exists(designation, save_dir):
        raise ValueError(f"Обозначение {designation} уже занято — такой файл уже есть в папке")


def suggest_designation(level, product=None, module=None, save_dir=None):
    """level: 'product' | 'module' | 'part'. Scans save_dir for existing designations and
    proposes the next free number at the requested level; caller may still edit it by hand."""
    existing = _scan_designations(save_dir)

    if level == "product":
        used = [a for (a, b, c) in existing]
        return f"{(max(used) + 1) if used else 1:04d}.0000.0000"

    if product is None:
        raise ValueError("Нужно указать номер изделия")
    product = int(product)

    if level == "module":
        used = [b for (a, b, c) in existing if a == product]
        return f"{product:04d}.{(max(used) + 1) if used else 1:04d}.0000"

    if level == "part":
        mod = int(module) if module else 0
        used = [c for (a, b, c) in existing if a == product and b == mod]
        return f"{product:04d}.{mod:04d}.{(max(used) + 1) if used else 1:04d}"

    raise ValueError(f"Неизвестный уровень: {level!r}")


def list_known_products(save_dir=None):
    """Distinct product numbers already used by files in save_dir — for the 'Номер изделия' picker."""
    return sorted({a for (a, b, c) in _scan_designations(save_dir)})


def list_known_modules(product, save_dir=None):
    """Distinct module numbers already used under a given product — for the 'Номер модуля' picker."""
    product = int(product)
    return sorted({b for (a, b, c) in _scan_designations(save_dir) if a == product and b != 0})


def _ensure_catalog():
    """Create catalog.xlsx with starter reference data if it doesn't exist yet next to the
    app (fresh install on a new machine) — the app should work out of the box, not require
    the user to build the spreadsheet themselves first."""
    if os.path.exists(CATALOG_PATH):
        return
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Трубы"
    ws.append(["Ширина_мм", "Высота_мм", "Толщина_стенки_мм", "Название"])
    for row in [
        (20, 20, 2, "20x20x2"), (25, 25, 2, "25x25x2"), (30, 30, 2, "30x30x2"),
        (40, 40, 2, "40x40x2"), (40, 40, 3, "40x40x3"), (50, 50, 2, "50x50x2"),
        (50, 50, 3, "50x50x3"), (60, 40, 3, "60x40x3"), (60, 60, 3, "60x60x3"),
        (80, 40, 3, "80x40x3"),
    ]:
        ws.append(list(row))

    ws2 = wb.create_sheet("Листы")
    ws2.append(["Толщина_мм"])
    for t in [1, 1.5, 2, 3, 4, 5, 6, 8, 10]:
        ws2.append([t])

    ws3 = wb.create_sheet("Наименования_деталей")
    ws3.append(["Наименование"])
    for name in ["Стойка", "Перекладина", "Основание", "Кронштейн", "Опора",
                 "Ручка", "Ступень", "Раскос", "Косынка", "Заглушка", "Пластина крепёжная"]:
        ws3.append([name])

    ws4 = wb.create_sheet("Крепёж")
    ws4.append(["Обозначение", "Диаметр_мм"])
    for d in ["M6", "M8", "M10", "M12", "M16"]:
        ws4.append([d, int(d[1:])])

    ws5 = wb.create_sheet("Цвета_покрытия")
    ws5.append(["RAL", "Название"])
    for ral, name in [
        ("RAL 9005", "Чёрный матовый"), ("RAL 9010", "Белый"), ("RAL 5010", "Синий"),
        ("RAL 6005", "Зелёный мох"), ("RAL 3020", "Красный"), ("RAL 1023", "Жёлтый"),
        ("RAL 7040", "Серый окна"),
    ]:
        ws5.append([ral, name])

    ws6 = wb.create_sheet("Профили_прокат")
    ws6.append(["Тип", "Обозначение", "Комментарий"])
    for row in [
        ("Уголок", "50x50x5", ""), ("Уголок", "63x63x6", ""),
        ("Швеллер", "8У", ""), ("Швеллер", "10У", ""),
        ("Труба круглая", "32x2", "не создаётся кнопкой пока — только справочно"),
        ("Труба круглая", "42x2", "не создаётся кнопкой пока — только справочно"),
    ]:
        ws6.append(list(row))

    wb.save(CATALOG_PATH)


def open_catalog():
    """Open catalog.xlsx directly in Excel for editing (bulk copy/paste, etc.)."""
    _ensure_catalog()
    os.startfile(CATALOG_PATH)
    return {"path": CATALOG_PATH}


def _read_catalog_sheet(sheet_name):
    from openpyxl import load_workbook
    _ensure_catalog()
    wb = load_workbook(CATALOG_PATH, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            rows.append(row)
    return rows


def list_tube_presets():
    """Standard tube cross-sections from catalog.xlsx ('Трубы' sheet) for the preset dropdown."""
    out = []
    for row in _read_catalog_sheet("Трубы"):
        width, height, wall = row[0], row[1], row[2]
        name = row[3] if len(row) > 3 and row[3] else f"{width}x{height}x{wall}"
        out.append({"width_mm": width, "height_mm": height, "wall_mm": wall, "name": name})
    return out


def list_sheet_thicknesses():
    """Standard sheet thicknesses from catalog.xlsx ('Листы' sheet) for the preset dropdown."""
    return [row[0] for row in _read_catalog_sheet("Листы")]


def list_part_name_templates():
    """Common part-name patterns from catalog.xlsx ('Наименования_деталей') for Description autocomplete."""
    return [row[0] for row in _read_catalog_sheet("Наименования_деталей")]


def create_assembly(designation, description="", save_dir=None):
    """Empty assembly container — used for 'Изделие' (a.0000.0000) and 'Модуль' (a.b.0000)."""
    _validate_designation(designation)
    save_dir = save_dir or DEFAULT_SAVE_DIR
    _check_not_duplicate(designation, save_dir)

    app = connect()
    doc = app.Documents.Add(kAssemblyDocumentObject, ASSEMBLY_TEMPLATE, True)
    _apply_pdm_properties(doc, designation, description, None)

    path = os.path.join(save_dir, designation + ".iam")
    doc.SaveAs(path, False)
    return {"path": path, "designation": designation}


def list_materials():
    """Materials available in the default part template's library."""
    app = connect()
    doc = app.Documents.Add(kPartDocumentObject, PART_TEMPLATE, False)
    try:
        return [m.Name for m in doc.Materials]
    finally:
        doc.Close(False)


def _corner_of(rect, x, y):
    """Find the SketchPoint of a rectangle's lines at (x, y) in sketch-space (cm)."""
    for i in range(1, rect.Count + 1):
        line = rect.Item(i)
        for pt in (line.StartSketchPoint, line.EndSketchPoint):
            if abs(pt.Geometry.X - x) < 1e-6 and abs(pt.Geometry.Y - y) < 1e-6:
                return pt
    raise RuntimeError(f"Sketch corner not found at ({x}, {y})")


def _fully_constrain_rectangle(sketch, tg, rect, w, h, width_expr, height_expr, ground=True):
    """Ground the (0,0) corner and dimension the rectangle's own width/height so the sketch
    is fully constrained (not just drawn at coordinates that happen to match). Returns the
    (origin, right, top) corner points for further use (e.g. positioning a nested rectangle)."""
    origin_pt = _corner_of(rect, 0, 0)
    right_pt = _corner_of(rect, w, 0)
    top_pt = _corner_of(rect, 0, h)
    if ground:
        sketch.GeometricConstraints.AddGround(origin_pt)
    dimW = sketch.DimensionConstraints.AddTwoPointDistance(
        origin_pt, right_pt, kHorizontalDim, tg.CreatePoint2d(w / 2, -0.5), False)
    dimH = sketch.DimensionConstraints.AddTwoPointDistance(
        origin_pt, top_pt, kVerticalDim, tg.CreatePoint2d(-0.5, h / 2), False)
    dimW.Parameter.Expression = width_expr
    dimH.Parameter.Expression = height_expr
    return origin_pt, right_pt, top_pt


def create_sheet_part(designation, length_mm, width_mm, thickness_mm,
                       description="", material_name=None, save_dir=None):
    """Flat rectangular plate — the 'Лист' template. Dimensions in mm."""
    _validate_designation(designation)
    save_dir = save_dir or DEFAULT_SAVE_DIR
    _check_not_duplicate(designation, save_dir)
    app = connect()
    tg = app.TransientGeometry
    doc = app.Documents.Add(kPartDocumentObject, PART_TEMPLATE, True)
    compDef = doc.ComponentDefinition

    length_cm, width_cm, thickness_cm = _mm(length_mm), _mm(width_mm), _mm(thickness_mm)
    params = compDef.Parameters.UserParameters
    params.AddByExpression("Length", f"{length_mm} mm", "mm")
    params.AddByExpression("Width", f"{width_mm} mm", "mm")
    params.AddByExpression("Thickness", f"{thickness_mm} mm", "mm")

    xy = compDef.WorkPlanes.Item(3)
    sketch = compDef.Sketches.Add(xy)
    rect = sketch.SketchLines.AddAsTwoPointRectangle(
        tg.CreatePoint2d(0, 0), tg.CreatePoint2d(length_cm, width_cm))
    _fully_constrain_rectangle(sketch, tg, rect, length_cm, width_cm, "Length", "Width")

    profile = sketch.Profiles.AddForSolid()
    extrudeDef = compDef.Features.ExtrudeFeatures.CreateExtrudeDefinition(profile, kNewBodyOperation)
    extrudeDef.SetDistanceExtent(thickness_cm, kPositiveExtentDirection)
    compDef.Features.ExtrudeFeatures.Add(extrudeDef)
    doc.Update()

    _apply_pdm_properties(doc, designation, description or "Лист", material_name)

    path = os.path.join(save_dir, designation + ".ipt")
    doc.SaveAs(path, False)
    return {"path": path, "designation": designation}


def create_tube_part(designation, outer_width_mm, outer_height_mm, wall_mm, length_mm,
                      description="", material_name=None, save_dir=None):
    """Hollow rectangular/square tube (profile = two nested rectangles) — the 'Труба' template.
    Dimensions in mm. For a square tube pass outer_width_mm == outer_height_mm."""
    _validate_designation(designation)
    save_dir = save_dir or DEFAULT_SAVE_DIR
    _check_not_duplicate(designation, save_dir)
    app = connect()
    tg = app.TransientGeometry
    doc = app.Documents.Add(kPartDocumentObject, PART_TEMPLATE, True)
    compDef = doc.ComponentDefinition

    w, h, t, length = (_mm(outer_width_mm), _mm(outer_height_mm), _mm(wall_mm), _mm(length_mm))
    if t * 2 >= min(w, h):
        raise ValueError("Толщина стенки не может быть больше половины меньшей стороны профиля")

    params = compDef.Parameters.UserParameters
    params.AddByExpression("OuterWidth", f"{outer_width_mm} mm", "mm")
    params.AddByExpression("OuterHeight", f"{outer_height_mm} mm", "mm")
    params.AddByExpression("WallThickness", f"{wall_mm} mm", "mm")
    params.AddByExpression("Length", f"{length_mm} mm", "mm")

    xy = compDef.WorkPlanes.Item(3)
    sketch = compDef.Sketches.Add(xy)
    # corner-anchored (not centered) so both loops can be pinned to the same origin corner
    outerRect = sketch.SketchLines.AddAsTwoPointRectangle(
        tg.CreatePoint2d(0, 0), tg.CreatePoint2d(w, h))
    innerRect = sketch.SketchLines.AddAsTwoPointRectangle(
        tg.CreatePoint2d(t, t), tg.CreatePoint2d(w - t, h - t))

    outer_origin, _, _ = _fully_constrain_rectangle(
        sketch, tg, outerRect, w, h, "OuterWidth", "OuterHeight")

    inner_origin = _corner_of(innerRect, t, t)
    inner_far = _corner_of(innerRect, w - t, h - t)
    # position the inner loop's near corner relative to the outer origin (both = WallThickness)
    dimPosX = sketch.DimensionConstraints.AddTwoPointDistance(
        outer_origin, inner_origin, kHorizontalDim, tg.CreatePoint2d(t / 2, -0.5), False)
    dimPosY = sketch.DimensionConstraints.AddTwoPointDistance(
        outer_origin, inner_origin, kVerticalDim, tg.CreatePoint2d(-0.5, t / 2), False)
    dimPosX.Parameter.Expression = "WallThickness"
    dimPosY.Parameter.Expression = "WallThickness"
    # size the inner loop itself, derived from the outer dimensions minus the wall
    dimInnerW = sketch.DimensionConstraints.AddTwoPointDistance(
        inner_origin, inner_far, kHorizontalDim, tg.CreatePoint2d((w) / 2, t / 2), False)
    dimInnerH = sketch.DimensionConstraints.AddTwoPointDistance(
        inner_origin, inner_far, kVerticalDim, tg.CreatePoint2d(t / 2, h / 2), False)
    dimInnerW.Parameter.Expression = "OuterWidth - 2 * WallThickness"
    dimInnerH.Parameter.Expression = "OuterHeight - 2 * WallThickness"

    profile = sketch.Profiles.AddForSolid()
    extrudeDef = compDef.Features.ExtrudeFeatures.CreateExtrudeDefinition(profile, kNewBodyOperation)
    extrudeDef.SetDistanceExtent(length, kPositiveExtentDirection)
    compDef.Features.ExtrudeFeatures.Add(extrudeDef)
    doc.Update()

    _apply_pdm_properties(
        doc, designation,
        description or f"Труба {outer_width_mm}x{outer_height_mm}x{wall_mm}",
        material_name)

    path = os.path.join(save_dir, designation + ".ipt")
    doc.SaveAs(path, False)
    return {"path": path, "designation": designation}
